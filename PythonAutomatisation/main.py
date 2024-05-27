import argparse
import glob
import os
import pathlib
import shutil
import subprocess
from math import floor
import re

import xlsxwriter
from openpyxl.reader.excel import load_workbook

import fshandler
from ennemy import Ennemy
from gamedata import GameData

########################################################################
# Const
########################################################################
FOLDER_INPUT = "OriginalFiles"
FOLDER_OUTPUT = "OutputFiles"
FILE_INPUT_BATTLE = os.path.join(FOLDER_INPUT, "decompressed_battle")  # Path for battle.fs decompressed
FILE_OUTPUT_BATTLE = os.path.join(FOLDER_OUTPUT, "battle")
FILE_BATTLE_SPECIAL_PATH_FORMAT = os.path.join("fre", "battle")
FILE_MONSTER_INPUT_PATH = os.path.join(FILE_INPUT_BATTLE, FILE_BATTLE_SPECIAL_PATH_FORMAT)  # Final path for .dat file (dependent of deling-CLI.exe)
FILE_MONSTER_OUTPUT_PATH = os.path.join(FILE_OUTPUT_BATTLE, FILE_BATTLE_SPECIAL_PATH_FORMAT)
FILE_MONSTER_INPUT_REGEX = os.path.join(FILE_MONSTER_INPUT_PATH, "c0m*.dat")
FILE_MONSTER_OUTPUT_REGEX = os.path.join(FILE_MONSTER_OUTPUT_PATH, "c0m*.dat")
FILE_XLSX = os.path.join(FOLDER_OUTPUT, "ifrit.xlsx")

COL_MONSTER_INFO = 0
COL_STAT = 3
COL_DEF = 14
COL_ITEM = 17
COL_MISC = 25
COL_ABILITIES = 28
ROW_FILE_DATA = 42
COL_FILE_DATA = 0
ROW_MAG = 1
ROW_MUG = 5
ROW_DROP = 9
ROW_MONSTER_NAME = 1
ROW_MONSTER_LVL = 2
ROW_MONSTER_COMBAT_TEXT = 3

COL_SHEET_LOW_LVL = 1
COL_SHEET_MED_LVL = 3
COL_SHEET_HIGH_LVL = 5
DEFAULT_MONSTER_LVL = 10
COL_GRAPH_PER_LVL = 4
ROW_GRAPH_PER_LVL = 33
NB_MAX_ABILITIES = 16

STAT_GRAPH_CELL_PLACEMENT = 'N34'
STAT_GRAPH_WIDTH = 900
STAT_GRAPH_HEIGHT = 500

REF_DATA_COL_ABILITIES_TYPE = 0
REF_DATA_COL_ABILITIES = 1
REF_DATA_COL_MAGIC = 2
REF_DATA_COL_ITEM = 3
REF_DATA_SHEET_TITLE = 'ref_data'

MAX_COMBAT_TXT = 8
MAX_SHEET_TITLE_SIZE = 31
INVALID_CHAR_TITLE_EXCEL_LIST = ['[',']', ':', '*', '?', '/', '\\']
########################################################################
# Useful
########################################################################

def create_checksum_file(ennemy_list, file_name):
    with open(os.path.join(FOLDER_OUTPUT, file_name), "w") as f:
        for key, ennemy in ennemy_list.items():
            f.write("File name: {} | Checksum: {}\n".format(ennemy.origin_file_name, ennemy.origin_file_checksum))


########################################################################
# Export from dat to XLSX
########################################################################
def export_to_xlsx(ennemy_list, game_data: Ennemy()):
    workbook = xlsxwriter.Workbook("OutputFiles/ifrit.xlsx")

    # Style creation
    column_title_style = workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#b2b2b2', 'align': 'center'})
    border_style = workbook.add_format({'border': 1})
    row_title_style = workbook.add_format({'border': 1, 'bold': True})
    magic_style = workbook.add_format({'border': 1, 'bg_color': '#b8cce4'})
    status_style = workbook.add_format({'border': 1, 'bg_color': '#b9b085'})
    drop_style = workbook.add_format({'border': 1, 'bg_color': '#ccc0da'})
    mug_style = workbook.add_format({'border': 1, 'bg_color': '#7aeca0'})
    title_magic_style = workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#b8cce4'})
    title_status_style = workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#b9b085'})
    title_drop_style = workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#ccc0da'})
    title_mug_style = workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#7aeca0'})
    percent_style = workbook.add_format({'num_format': '#,##0.00%', 'border': 1})

    # Chart
    ## Stat chart

    chart_stat = {}
    tab_list = []
    for key, ennemy in ennemy_list.items():

        # Tab (sheet) creation. Checking name doesn't already exist
        file_index = int(re.search(r'\d{3}', ennemy.origin_file_name).group())
        file_name = str(file_index) + " - " + ennemy.name
        if file_name == '':
            file_name = "Empty"
        while file_name in tab_list:
            file_name += " dub"

        if len(file_name) > MAX_SHEET_TITLE_SIZE:
            file_name = file_name[:MAX_SHEET_TITLE_SIZE]
        for char in INVALID_CHAR_TITLE_EXCEL_LIST:
            file_name = file_name.replace(char, ';')
        worksheet = workbook.add_worksheet(file_name)
        chart_stat[ennemy] = workbook.add_chart({'type': 'line'})
        tab_list.append(file_name)

        # Column position of different "menu"
        column_index = {}
        column_index['stat'] = COL_STAT
        column_index['def'] = COL_DEF
        column_index['item'] = COL_ITEM
        column_index['misc'] = COL_MISC
        column_index['graph_stat'] = COL_GRAPH_PER_LVL + 1
        column_index['abilities'] = COL_ABILITIES

        # Titles
        worksheet.write_row(0, COL_MONSTER_INFO, ["Monster info"], cell_format=column_title_style)
        worksheet.write_row(0, column_index['stat'] + 1, ["Value 1", "Value 2", "Value 3", "Value 4", "Impact 1", "Impact 2", "Impact 3", "Impact 4", "Total"],
                            cell_format=column_title_style)
        worksheet.write_row(0, column_index['def'], ["Defense name", "% Resistance"], cell_format=column_title_style)
        worksheet.write_row(0, column_index['item'] + 1, ["Low Level", "Number", "Medium Level", "Number", "High Level", "Number"],
                            cell_format=column_title_style)
        worksheet.write_row(0, column_index['misc'], ["Property name", "Value"], cell_format=column_title_style)
        worksheet.merge_range(xlsxwriter.utility.xl_col_to_name(COL_ABILITIES) + "1:" + xlsxwriter.utility.xl_col_to_name(COL_ABILITIES + 2) + "1", "Low Level",
                              cell_format=column_title_style)
        worksheet.merge_range(xlsxwriter.utility.xl_col_to_name(COL_ABILITIES + 3) + "1:" + xlsxwriter.utility.xl_col_to_name(COL_ABILITIES + 5) + "1",
                              "Medium Level", cell_format=column_title_style)
        worksheet.merge_range(xlsxwriter.utility.xl_col_to_name(COL_ABILITIES + 6) + "1:" + xlsxwriter.utility.xl_col_to_name(COL_ABILITIES + 8) + "1",
                              "High Level", cell_format=column_title_style)
        worksheet.write_row(1, COL_ABILITIES, ["Type", "Ability", "Animation", "Type", "Ability", "Animation", "Type", "Ability", "Animation"],
                            cell_format=column_title_style)

        # File info not link to the monster data
        worksheet.write(ROW_FILE_DATA, COL_FILE_DATA, "File data", column_title_style)
        worksheet.write(ROW_FILE_DATA + 1, COL_FILE_DATA, "Original file name", row_title_style)
        worksheet.write(ROW_FILE_DATA + 1, COL_FILE_DATA + 1, key, border_style)

        # Graph level
        worksheet.write(ROW_GRAPH_PER_LVL, COL_GRAPH_PER_LVL, "Level", column_title_style)
        for i in range(1, 101):
            worksheet.write(ROW_GRAPH_PER_LVL + i, COL_GRAPH_PER_LVL, i, row_title_style)

        # Index setting
        row_index = {}
        row_index['stat'] = 1
        row_index['stat'] = 1
        row_index['def'] = 1
        row_index['item'] = 1
        row_index['misc'] = 1
        row_index['abilities'] = 2
        index = {}
        index['elem_def'] = 0
        index['status_def'] = 0

        # General Monster info
        worksheet.write(ROW_MONSTER_LVL, COL_MONSTER_INFO, "Monster LVL", row_title_style)
        worksheet.write(ROW_MONSTER_LVL, COL_MONSTER_INFO + 1, DEFAULT_MONSTER_LVL, border_style)
        worksheet.write(ROW_MONSTER_NAME, COL_MONSTER_INFO, "Name", row_title_style)
        worksheet.write(ROW_MONSTER_NAME, COL_MONSTER_INFO + 1, ennemy.name, border_style)

        for i in range(len(ennemy.text_battle)):
            worksheet.write(ROW_MONSTER_COMBAT_TEXT + i, COL_MONSTER_INFO, "Combat text {}".format(i), row_title_style)
            worksheet.write(ROW_MONSTER_COMBAT_TEXT + i, COL_MONSTER_INFO + 1, "{}".format(ennemy.text_battle[i]), border_style)

        # Filling the Excel
        for el in ennemy.data:
            try:
                # Stat menu
                column_index['stat'] = COL_STAT
                impact = []
                if el['name'] in game_data.stat_values:
                    # Writing title of row
                    worksheet.write(row_index['stat'], COL_STAT, el['pretty_name'], row_title_style)
                    column_index['stat'] += 1
                    # Writing the 4 values of stats
                    for el2 in el['value']:
                        worksheet.write(row_index['stat'], column_index['stat'], el2, border_style)
                        column_index['stat'] += 1

                    # Writing the Impact and total stat
                    monster_lvl_cell = xlsxwriter.utility.xl_col_to_name(COL_MONSTER_INFO + 1) + str(ROW_MONSTER_LVL + 1)
                    stat_cell = [None] * 4
                    # Title lvl column stat
                    worksheet.write(ROW_GRAPH_PER_LVL, column_index['graph_stat'], el['pretty_name'], column_title_style)
                    if el['name'] == 'hp':
                        # Impact 1
                        stat_cell[0] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 1) + str(row_index['stat'] + 1)
                        worksheet.write(row_index['stat'], column_index['stat'],
                                        '=FLOOR({}*({}*{}/20+{}),1)'.format(stat_cell[0], monster_lvl_cell, monster_lvl_cell, monster_lvl_cell), border_style)
                        # Impact 2
                        stat_cell[1] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 2) + str(row_index['stat'] + 1)
                        worksheet.write(row_index['stat'], column_index['stat'] + 1, '=10*{}'.format(stat_cell[1]), border_style)
                        # Impact 3
                        stat_cell[2] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 3) + str(row_index['stat'] + 1)
                        worksheet.write(row_index['stat'], column_index['stat'] + 2, '={}*100*{}'.format(stat_cell[2], monster_lvl_cell), border_style)
                        # Impact 4
                        stat_cell[3] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 4) + str(row_index['stat'] + 1)
                        worksheet.write(row_index['stat'], column_index['stat'] + 3, '=1000*{}'.format(stat_cell[3]), border_style)
                        # Total Sum
                        worksheet.write(row_index['stat'], column_index['stat'] + 4,
                                        '=SUM({}:{})'.format(xlsxwriter.utility.xl_col_to_name(COL_STAT + 5) + str(row_index['stat'] + 1),
                                                             xlsxwriter.utility.xl_col_to_name(COL_STAT + 8) + str(row_index['stat'] + 1)), border_style)
                        # Total Formula
                        # The HP is harder to show on a graph, so we divide it by 100 to show it better
                        worksheet.write(ROW_GRAPH_PER_LVL, column_index['graph_stat'], el['pretty_name'], column_title_style)
                        for i in range(1, 101):
                            monster_lvl_cell = xlsxwriter.utility.xl_col_to_name(COL_GRAPH_PER_LVL) + str(ROW_GRAPH_PER_LVL + i + 1)
                            # i corresponding to the monster level
                            # /10 because HP is too high to be plot correctly with others values.
                            str_formula = '=(FLOOR({}*({}*{}/20+{}),1) + 10*{} + {}*100*{} + 1000*{})/100'.format(stat_cell[0], monster_lvl_cell,
                                                                                                                  monster_lvl_cell, monster_lvl_cell,
                                                                                                                  stat_cell[1], stat_cell[2], monster_lvl_cell,
                                                                                                                  stat_cell[3])
                            worksheet.write(ROW_GRAPH_PER_LVL + i, column_index['graph_stat'], str_formula, border_style)

                    elif el['name'] == 'str' or el['name'] == 'mag':
                        # Impact 1
                        stat_cell[0] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 1) + str(row_index['stat'] + 1)
                        worksheet.write(row_index['stat'], column_index['stat'], '=FLOOR({}*{}/40, 1)'.format(monster_lvl_cell, stat_cell[0]), border_style)
                        # Impact 2
                        stat_cell[1] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 2) + str(row_index['stat'] + 1)
                        worksheet.write(row_index['stat'], column_index['stat'] + 1, '=FLOOR({}/(4*{}),1)'.format(monster_lvl_cell, stat_cell[1]), border_style)
                        # Impact 3
                        stat_cell[2] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 3) + str(row_index['stat'] + 1)
                        worksheet.write(row_index['stat'], column_index['stat'] + 2, '=FLOOR({}/4,1)'.format(stat_cell[2]), border_style)
                        # Impact 4
                        stat_cell[3] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 4) + str(row_index['stat'] + 1)
                        worksheet.write(row_index['stat'], column_index['stat'] + 3,
                                        '=FLOOR({}*{}/(8*{}),1)'.format(monster_lvl_cell, monster_lvl_cell, stat_cell[3]), border_style)
                        # Total
                        worksheet.write(row_index['stat'], column_index['stat'] + 4,
                                        '=SUM({}:{})'.format(xlsxwriter.utility.xl_col_to_name(COL_STAT + 5) + str(row_index['stat'] + 1),
                                                             xlsxwriter.utility.xl_col_to_name(COL_STAT + 8) + str(row_index['stat'] + 1)), border_style)

                        # Total Formula
                        for i in range(1, 101):
                            monster_lvl_cell = xlsxwriter.utility.xl_col_to_name(COL_GRAPH_PER_LVL) + str(ROW_GRAPH_PER_LVL + i + 1)
                            # i corresponding to the monster level
                            str_formula = '=FLOOR({}*{}/40, 1)+FLOOR({}/(4*{}),1)+FLOOR({}/4,1)+FLOOR({}*{}/(8*{}),1)'.format(monster_lvl_cell, stat_cell[0],
                                                                                                                              monster_lvl_cell, stat_cell[1],
                                                                                                                              stat_cell[2], monster_lvl_cell,
                                                                                                                              monster_lvl_cell, stat_cell[3])
                            worksheet.write(ROW_GRAPH_PER_LVL + i, column_index['graph_stat'], str_formula, border_style)
                    elif el['name'] == 'vit' or el['name'] == 'spr' or el['name'] == 'spd' or el['name'] == 'eva':
                        # Impact 1
                        stat_cell[0] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 1) + str(row_index['stat'] + 1)
                        worksheet.write(row_index['stat'], column_index['stat'], '={}*{}'.format(monster_lvl_cell, stat_cell[0]), border_style)
                        # Impact 2
                        stat_cell[1] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 2) + str(row_index['stat'] + 1)
                        worksheet.write(row_index['stat'], column_index['stat'] + 1, '=FLOOR({}/{},1)'.format(monster_lvl_cell, stat_cell[1]), border_style)
                        # Impact 3
                        stat_cell[2] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 3) + str(row_index['stat'] + 1)
                        worksheet.write(row_index['stat'], column_index['stat'] + 2, '={}'.format(stat_cell[2]), border_style)
                        # Impact 4
                        stat_cell[3] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 4) + str(row_index['stat'] + 1)
                        worksheet.write(row_index['stat'], column_index['stat'] + 3, '=-FLOOR({}/{},1)'.format(monster_lvl_cell, stat_cell[3]), border_style)
                        # Total
                        worksheet.write(row_index['stat'], column_index['stat'] + 4,
                                        '=SUM({}:{})'.format(xlsxwriter.utility.xl_col_to_name(COL_STAT + 5) + str(row_index['stat'] + 1),
                                                             xlsxwriter.utility.xl_col_to_name(COL_STAT + 8) + str(row_index['stat'] + 1)), border_style)
                        # Total Formula
                        for i in range(1, 101):
                            monster_lvl_cell = xlsxwriter.utility.xl_col_to_name(COL_GRAPH_PER_LVL) + str(ROW_GRAPH_PER_LVL + i + 1)
                            # i corresponding to the monster level
                            str_formula = '={}*{}+FLOOR({}/{},1)+{}-FLOOR({}/{},1)'.format(monster_lvl_cell, stat_cell[0], monster_lvl_cell, stat_cell[1],
                                                                                           stat_cell[2], monster_lvl_cell, stat_cell[3])
                            worksheet.write(ROW_GRAPH_PER_LVL + i, column_index['graph_stat'], str_formula, border_style)

                    row_index['stat'] += 1
                    column_index['graph_stat'] += 1

                    # Creating chart with data computed
                    current_stat_column_str = xlsxwriter.utility.xl_col_to_name(column_index['graph_stat'] - 1)
                    lvl_range_str = '=\'' + file_name + '\'!$E$35:$E$135'
                    lvl_stat_str = '=\'' + file_name + '\'!${}$35:${}$135'.format(current_stat_column_str, current_stat_column_str)
                    graph_serie_name = el['pretty_name']
                    if el['name'] == 'hp':
                        graph_serie_name += '/100'
                    chart_stat[ennemy].add_series({'name': graph_serie_name, 'categories': lvl_range_str, 'values': lvl_stat_str, 'smooth': True})
                # Def menu
                elif el['name'] in ['elem_def', 'status_def']:
                    for el2 in el['value']:
                        if el['name'] == 'elem_def':
                            worksheet.write(row_index['def'], column_index['def'], game_data.magic_type_values[index[el['name']]], title_magic_style)
                            worksheet.write(row_index['def'], column_index['def'] + 1, el2, magic_style)
                        elif el['name'] == 'status_def':
                            worksheet.write(row_index['def'], column_index['def'], game_data.status_values[index[el['name']]], title_status_style)
                            worksheet.write(row_index['def'], column_index['def'] + 1, el2, status_style)
                        index[el['name']] += 1
                        row_index['def'] += 1
                # Item menu (containing draw too)
                elif el['name'] in ['low_lvl_mag', 'med_lvl_mag', 'high_lvl_mag', 'low_lvl_mug', 'med_lvl_mug', 'high_lvl_mug', 'low_lvl_drop', 'med_lvl_drop',
                                    'high_lvl_drop']:  # Items
                    if 'mag' in el['name']:
                        row_index['item'] = ROW_MAG
                    elif 'mug' in el['name']:
                        row_index['item'] = ROW_MUG
                    elif 'drop' in el['name']:
                        row_index['item'] = ROW_DROP
                    # Indew as we fill the data first per 'type' (draw, mug, drop) then per column. And there is 3 column.
                    index['draw'] = 1
                    index['mug'] = 1
                    index['drop'] = 1
                    # Going through value
                    for el2 in el['value']:
                        if 'low' in el['name']:
                            col_index = column_index['item'] + COL_SHEET_LOW_LVL
                        elif 'med' in el['name']:
                            col_index = column_index['item'] + COL_SHEET_MED_LVL
                        elif 'high' in el['name']:
                            col_index = column_index['item'] + COL_SHEET_HIGH_LVL
                        else:  # Should not happen
                            print("Problem on column index")
                            col_index = column_index['item'] + COL_SHEET_HIGH_LVL + 2
                        if 'mag' in el['name']:
                            worksheet.write(row_index['item'], column_index['item'], "Draw {}".format(index['draw']), title_magic_style)
                            index['draw'] += 1
                            worksheet.write(row_index['item'], col_index, game_data.magic_values[el2['ID']]['ref'], magic_style)
                            worksheet.write(row_index['item'], col_index + 1, el2['value'], magic_style)
                        elif 'mug' in el['name']:
                            worksheet.write(row_index['item'], column_index['item'], "Mug {}".format(index['mug']), title_mug_style)
                            index['mug'] += 1
                            worksheet.write(row_index['item'], col_index, game_data.item_values[el2['ID']]['ref'], mug_style)
                            worksheet.write(row_index['item'], col_index + 1, el2['value'], mug_style)
                        elif 'drop' in el['name']:
                            worksheet.write(row_index['item'], column_index['item'], "Drop {}".format(index['drop']), title_drop_style)
                            index['drop'] += 1
                            worksheet.write(row_index['item'], col_index, game_data.item_values[el2['ID']]['ref'], drop_style)
                            worksheet.write(row_index['item'], col_index + 1, el2['value'], drop_style)
                        row_index['item'] += 1
                # Misc menu
                elif el['name'] in game_data.MISC_ORDER:
                    worksheet.write(row_index['misc'], column_index['misc'], el['pretty_name'], row_title_style)
                    if el['name'] == "mug_rate" or el['name'] == "drop_rate":  # Percent style need to divide by 100
                        worksheet.write(row_index['misc'], column_index['misc'] + 1, el['value'] / 100, percent_style)
                    else:
                        worksheet.write(row_index['misc'], column_index['misc'] + 1, floor(el['value']), border_style)
                    row_index['misc'] += 1
                # Abilities menu
                elif el['name'] in game_data.ABILITIES_HIGHNESS_ORDER:
                    for el2 in el['value']:
                        ability_type = game_data.ennemy_abilities_type_values[el2['type']]
                        if ability_type['name'] == "Magic":
                            ability_name = game_data.magic_values[el2['id']]['ref']
                            col_ab_str = xlsxwriter.utility.xl_col_to_name(REF_DATA_COL_MAGIC)
                        elif ability_type['name'] == "Custom":
                            ability_name = game_data.ennemy_abilities_values[el2['id']]['ref']
                            col_ab_str = xlsxwriter.utility.xl_col_to_name(REF_DATA_COL_ABILITIES)
                        elif ability_type['name'] == "Item":
                            ability_name = game_data.item_values[el2['id']]['ref']
                            col_ab_str = xlsxwriter.utility.xl_col_to_name(REF_DATA_COL_ITEM)
                        else:
                            if el2['id'] < len(game_data.ennemy_abilities_values):
                                ability_name = game_data.ennemy_abilities_values[el2['id']]['ref']
                            else:
                                game_data.ennemy_abilities_values[el2['id']] = {'name':"Temp Garbage", 'ref':str(el2['id']) + ":Temp Garbage"}
                                ability_name = game_data.ennemy_abilities_values[el2['id']]['ref']
                            col_ab_str = xlsxwriter.utility.xl_col_to_name(REF_DATA_COL_ABILITIES)

                        worksheet.write(row_index['abilities'], column_index['abilities'], ability_type['ref'], border_style)
                        # Excel data validation
                        col_str = xlsxwriter.utility.xl_col_to_name(REF_DATA_COL_ABILITIES_TYPE)
                        source_str = '=' + REF_DATA_SHEET_TITLE + '!$' + col_str + '1:$' + col_str + '$245'
                        worksheet.data_validation(row_index['abilities'], column_index['abilities'], row_index['abilities'], column_index['abilities'],
                                                  {'validate': 'list', 'source': source_str})
                        worksheet.write(row_index['abilities'], column_index['abilities'] + 1, ability_name, border_style)
                        source_str = '=' + REF_DATA_SHEET_TITLE + '!$' + col_ab_str + '1:$' + col_ab_str + '$384'
                        worksheet.data_validation(row_index['abilities'], column_index['abilities'] + 1, row_index['abilities'], column_index['abilities'] + 1,
                                                  {'validate': 'list', 'source': source_str})

                        worksheet.write(row_index['abilities'], column_index['abilities'] + 2, el2['animation'], border_style)
                        row_index['abilities'] += 1
                    row_index['abilities'] = 2
                    column_index['abilities'] += 3
            except IndexError:
                raise IndexError("Unknown error on file {} for monster name {}".format(key, ennemy.name))

        # Chart management
        chart_stat[ennemy].set_title({'name': 'Stat graph'})
        chart_stat[ennemy].set_x_axis({'name': 'Level'})
        chart_stat[ennemy].set_y_axis({'name': 'Stat'})
        chart_stat[ennemy].set_size({'width': STAT_GRAPH_WIDTH, 'height': STAT_GRAPH_HEIGHT})

        worksheet.insert_chart(STAT_GRAPH_CELL_PLACEMENT, chart_stat[ennemy])

        worksheet.autofit()

    # Creating reference data on last tab
    worksheet = workbook.add_worksheet(REF_DATA_SHEET_TITLE)
    for index, el in game_data.ennemy_abilities_type_values.items():
        worksheet.write(index, REF_DATA_COL_ABILITIES_TYPE, el['ref'])
    for index, el in game_data.ennemy_abilities_values.items():
        worksheet.write(index, REF_DATA_COL_ABILITIES, el['ref'])
    for index, el in game_data.magic_values.items():
        worksheet.write(index, REF_DATA_COL_MAGIC, el['ref'])
    for index, el in game_data.item_values.items():
        worksheet.write(index, REF_DATA_COL_ITEM, el['ref'])
    worksheet.autofit()
    workbook.close()


def dat_to_xlsx(file_list):
    monster = {}

    print("Getting game data")
    game_data = GameData()
    game_data.load_all()

    print("Reading ennemy files")
    for file in file_list:
        file_name = os.path.basename(file)
        monster[file_name] = Ennemy()
        monster[file_name].load_file_data(file, game_data)
        monster[file_name].load_name(game_data)

    print("Analysing ennemy files")
    for file in file_list:
        file_name = os.path.basename(file)
        monster[file_name].analyse_loaded_data(game_data)

    print("Creating checksum file")
    create_checksum_file(monster, "checksum_origin_file.txt")

    print("Writing to xlsx file")
    export_to_xlsx(monster, game_data)


########################################################################
# Import from XLSX to dat
########################################################################
def import_from_xlsx(file, ennemy, game_data):
    """
    As the module to write is different from the reading one, the one writing start at 0 for column and row, when this one, the reading, start at 1
    """
    wb = load_workbook(file)
    for sheet in wb:

        if sheet.title == REF_DATA_SHEET_TITLE:
            continue

        ennemy_origin_file = sheet.cell(ROW_FILE_DATA + 1 + 1, COL_FILE_DATA + 1 + 1).value
        ennemy[ennemy_origin_file] = Ennemy()
        current_ennemy = ennemy[ennemy_origin_file]
        current_ennemy.name = sheet.title
        ennemy[ennemy_origin_file].origin_file_name = ennemy_origin_file
        current_ennemy.data = [] # Reseting data as we need this var empty. It has been loaded to find others value like pointer
        # Stat reading
        row_index = 2
        for stat in game_data.stat_values:
            list_value = []
            for col_index in range(COL_STAT + 2, COL_STAT + 2 + 4):
                list_value.append(sheet.cell(row_index, col_index).value)
            current_ennemy.data.append({'name': stat, 'value': list_value})
            row_index += 1

        # Def reading
        list_value = []
        for i in range(2, len(game_data.magic_type_values) + 1):
            list_value.append(sheet.cell(i, COL_DEF + 1 + 1).value)
        current_ennemy.data.append({'name': 'elem_def', 'value': list_value})

        list_value = []
        for i in range(len(game_data.magic_type_values) + 1, len(game_data.magic_type_values) + len(game_data.status_values)):
            list_value.append(sheet.cell(i, COL_DEF + 1 + 1).value)
        current_ennemy.data.append({'name': 'status_def', 'value': list_value})

        # Item read
        item = ['mag', 'mug', 'drop']
        sub_item = ['low_lvl', 'med_lvl', 'high_lvl']
        row_index = 2
        col_index = COL_ITEM + 1 + 1
        list_value = []
        for el in item:
            for sub in sub_item:
                name = sub + "_" + el
                for i in range(4):
                    id_value = int(sheet.cell(row_index + i, col_index).value.split(':')[0])
                    value = sheet.cell(row_index + i, col_index + 1).value
                    list_value.append({'ID': id_value, 'value': value})
                current_ennemy.data.append({'name': name, 'value': list_value})
                list_value = []

                col_index += 2
            row_index += 4
            col_index = COL_ITEM + 1 + 1

        # Misc reading
        row_index = 2
        for misc in game_data.MISC_ORDER:
            value = sheet.cell(row_index, COL_MISC + 1 + 1).value
            if misc == "mug_rate" or misc == "drop_rate":
                value = value * 100
            current_ennemy.data.append({'name': misc, 'value': value})
            row_index += 1

        # Abilities reading
        col_index = COL_ABILITIES + 1
        for abilities in game_data.ABILITIES_HIGHNESS_ORDER:
            ability_set = []
            for i in range(3, NB_MAX_ABILITIES + 3):
                type = int(sheet.cell(i, col_index).value.split(':')[0])
                ability_id = int(sheet.cell(i, col_index+1).value.split(':')[0])
                animation = int(sheet.cell(i, col_index + 2).value)
                ability_set.append({'type': type, 'animation': animation, 'id': ability_id})
            current_ennemy.data.append({'name': abilities, 'value': ability_set})
            col_index += 3

        #Text reading
        combat_text_list = []
        for i in range(MAX_COMBAT_TXT):
            txt_value = sheet.cell(ROW_MONSTER_COMBAT_TEXT + 1+ i, 2).value
            if txt_value:
                combat_text_list.append(txt_value)
            else:
                break
        current_ennemy.data.append({'name': 'combat_text', 'value': combat_text_list})

def write_to_dat(ennemy_list, game_data: GameData, path: str):
    for key, ennemy in ennemy_list.items():
        ennemy.write_data_to_file(game_data, path)


def xlsx_to_dat(xlsx_file):
    ennemy = {}

    print("Getting game data")
    game_data = GameData()
    game_data.load_all()

    print("Importing data from xlsx")
    import_from_xlsx(xlsx_file, ennemy, game_data)

    print("Writing data to dat files")
    write_to_dat(ennemy, game_data, FILE_MONSTER_OUTPUT_PATH)

    print("Creating checksum file")
    create_checksum_file(ennemy, "checksum_output_file.txt")


if __name__ == "__main__":

    parser = argparse.ArgumentParser(prog="Ifrit Enhanced Command line",
                                     description="This program allow you to quickly translate your battle.fs in an easy to edit xlsx file, and transform it "
                                                 "back to battle.fs file")
    parser.add_argument("xlsx",
                        help="Define if you want to transform from battle.fs to xlsx or if you want to transform your modified xlsx back to a battle.fs",
                        choices=["fs_to_xlsx", "xlsx_to_fs", "both"])
    parser.add_argument("-d", "--delete",
                        help="Delete temporary file created (.dat extracted for example). Only applied for xlsx_to_fs command, as the temporary files are needed to build back to battle.fs",
                        action='store_true')
    parser.add_argument("-c", "--copy", help="Copy battle.fs to FF8 repository and to direct folder, expecting path to the FF8 folder", type=str)
    parser.add_argument("-cf", "--copyfile", help="Work only on file given", type=int)
    parser.add_argument("-l", "--limit", help="Limit to one monster only to test", action='store_true')
    parser.add_argument("-o", "--open", help="Open xlsx file for faster management", action='store_true')
    parser.add_argument("--nopack", help="Doesn't create a pack, only let intermediate file. Only applied to fs_to_xlsx", action='store_true')
    args = parser.parse_args()

    if args.xlsx == "fs_to_xlsx" or args.xlsx == "both":
        print("-------Unpacking fs file-------")
        fshandler.unpack(FOLDER_INPUT, FILE_INPUT_BATTLE)

        # Check if .dat exist
        file_monster = glob.glob(FILE_MONSTER_INPUT_REGEX)

        if not file_monster:
            raise FileNotFoundError("No .dat files found in {}".format(FILE_MONSTER_INPUT_PATH))
        if args.limit:
            file_monster = [file_monster[9]]  # Just to not work on all files everytime
        if args.copyfile:
            file_monster = [file_monster[args.copyfile]]  # Just to not work on all files everytime
        print("-------Transforming dat to XLSX-------")
        os.makedirs(FOLDER_OUTPUT, exist_ok=True)
        dat_to_xlsx(file_monster)

    if args.xlsx == "xlsx_to_fs" or args.xlsx == "both":
        print("-------Copying files from input to output-------")
        # Copying files from Input to Output before modifying
        os.makedirs(FILE_OUTPUT_BATTLE, exist_ok=True)
        os.makedirs(FILE_MONSTER_OUTPUT_PATH, exist_ok=True)
        shutil.copytree(FILE_MONSTER_INPUT_PATH, FILE_MONSTER_OUTPUT_PATH, dirs_exist_ok=True)

        print("-------Transforming XLSX to dat-------")
        xlsx_to_dat(FILE_XLSX)

        if not args.nopack:
            print("-------Packing to fs file-------")
            fshandler.pack(FILE_MONSTER_INPUT_PATH, FILE_OUTPUT_BATTLE)
        if args.delete:
            # Delete files
            print("-------Deleting files-------")
            ## Delete opened archive battle.fs (in Original file)
            shutil.rmtree(FILE_INPUT_BATTLE)
            ## Delete dat files in OutputFiles (in Output file)
            shutil.rmtree(os.path.join(FILE_OUTPUT_BATTLE, "fre"))

    if args.copy:
        for file in glob.glob(os.path.join(FILE_OUTPUT_BATTLE, "battle.*")):
            shutil.copy(file, args.copy)
        for file in glob.glob(os.path.join(FILE_MONSTER_OUTPUT_PATH, "*.*")):
            shutil.copy(file, os.path.join (args.copy, '..', '..', 'direct', 'battle'))
    if args.open:
        os.startfile(FILE_XLSX)
